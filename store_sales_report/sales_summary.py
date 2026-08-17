from __future__ import annotations

import argparse
import glob
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

import openpyxl
from dateutil import parser as dateparser
from openpyxl.styles import Alignment, Font, PatternFill


HEADER_ROW = 1
DATA_START_ROW = 2
SALES_COLUMN = 8  # Column H
NUMFMT = '₹#,##0.00'
EXPECTED_HEADERS = [
    "Date",
    "Bill No",
    "Customer Name",
    "Customer Mobile No",
    "Payment Method",
    "Sales Via",
    "Coupon Discount",
    "Total Rate(INR)",
]


@dataclass
class StoreSalesReport:
    store_name: str
    source_path: Path
    source_file: str
    total_sales: float
    period_start: datetime | None
    period_end: datetime | None


def normalize_text(value) -> str:
    return str(value or "").strip().replace(" ", "").lower()


def collect_files(file_args: list[str], output_name: str) -> list[Path]:
    if file_args:
        paths = [Path(arg).expanduser().resolve() for arg in file_args]
    else:
        paths = [Path(path).resolve() for path in glob.glob("*.xlsx")]

    result = []
    for path in paths:
        if not path.is_file():
            continue
        if path.suffix.lower() != ".xlsx":
            continue
        if path.name.startswith("~$"):
            continue
        if path.name == output_name:
            continue
        result.append(path)

    return sorted(result)


def validate_layout(ws, source_name: str) -> None:
    headers = [ws.cell(HEADER_ROW, col).value for col in range(1, len(EXPECTED_HEADERS) + 1)]
    normalized_headers = [normalize_text(value) for value in headers]
    normalized_expected = [normalize_text(value) for value in EXPECTED_HEADERS]

    if normalized_headers != normalized_expected:
        raise ValueError(
            f"{source_name}: unexpected header row. "
            f"Expected {EXPECTED_HEADERS}, found {headers}."
        )


def parse_any_date(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    text = str(value).strip()
    if not text:
        return None
    try:
        return dateparser.parse(text, dayfirst=True)
    except (ValueError, TypeError):
        return None


def to_number(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace(",", "").strip()
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def extract_store_report(path: Path) -> StoreSalesReport:
    workbook = openpyxl.load_workbook(path, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    validate_layout(worksheet, path.name)

    total = 0.0
    dates: list[datetime] = []
    for row_idx in range(DATA_START_ROW, worksheet.max_row + 1):
        value = to_number(worksheet.cell(row_idx, SALES_COLUMN).value)
        if value is not None:
            total += value
        date_value = parse_any_date(worksheet.cell(row_idx, 1).value)
        if date_value is not None:
            dates.append(date_value)

    period_start = min(dates) if dates else None
    period_end = max(dates) if dates else None

    return StoreSalesReport(
        store_name=path.stem,
        source_path=path,
        source_file=path.name,
        total_sales=total,
        period_start=period_start,
        period_end=period_end,
    )


def write_store_sheet(worksheet, report: StoreSalesReport) -> None:
    worksheet.column_dimensions["B"].width = 24
    worksheet.column_dimensions["C"].width = 18
    worksheet.column_dimensions["D"].width = 18

    worksheet["B2"] = f"{report.store_name} Sales Summary"
    worksheet["B2"].font = Font(name="Arial", size=14, bold=True, color="FF1B365D")
    worksheet["B3"] = f"Source file: {report.source_file}"
    worksheet["B3"].font = Font(name="Arial", size=10, color="FF555555")
    worksheet["B4"] = "Date From"
    worksheet["C4"] = report.period_start
    worksheet["B5"] = "Date To"
    worksheet["C5"] = report.period_end
    worksheet["B6"] = "Total Sales"
    worksheet["C6"] = report.total_sales

    for cell_ref in ("C4", "C5"):
        worksheet[cell_ref].number_format = "dd-mmm-yyyy"
    worksheet["C6"].number_format = NUMFMT

    for row in range(4, 7):
        worksheet[f"B{row}"].font = Font(name="Arial", size=11, bold=True)
        worksheet[f"C{row}"].font = Font(name="Arial", size=11)

    headers = ["Date", "Bill No", "Customer Name", "Customer Mobile No", "Payment Method",
               "Sales Via", "Coupon Discount", "Total Rate(INR)"]
    header_row = 8
    for col_idx, header in enumerate(headers, start=2):
        cell = worksheet.cell(header_row, col_idx, header)
        cell.font = Font(name="Arial", size=11, bold=True, color="FFFFFFFF")
        cell.fill = PatternFill(fill_type="solid", fgColor="FF1B365D")
        cell.alignment = Alignment(horizontal="center")

    workbook = openpyxl.load_workbook(report.source_path, data_only=True)
    source_ws = workbook[workbook.sheetnames[0]]
    row_out = header_row + 1
    for row_idx in range(DATA_START_ROW, source_ws.max_row + 1):
        row = [source_ws.cell(row_idx, col_idx).value for col_idx in range(1, 9)]
        if all(value is None for value in row):
            continue
        date_val = parse_any_date(row[0]) or row[0]
        sales_val = to_number(row[7])
        worksheet.cell(row_out, 2, date_val)
        worksheet.cell(row_out, 3, row[1])
        worksheet.cell(row_out, 4, row[2])
        worksheet.cell(row_out, 5, row[3])
        worksheet.cell(row_out, 6, row[4])
        worksheet.cell(row_out, 7, row[5])
        worksheet.cell(row_out, 8, row[6])
        worksheet.cell(row_out, 9, sales_val)
        worksheet.cell(row_out, 2).number_format = "dd-mmm-yyyy"
        worksheet.cell(row_out, 9).number_format = NUMFMT
        row_out += 1

    total_row = row_out + 1
    worksheet.cell(total_row, 8, "TOTAL")
    worksheet.cell(total_row, 9, report.total_sales)
    worksheet.cell(total_row, 9).number_format = NUMFMT
    for col_idx in range(8, 10):
        worksheet.cell(total_row, col_idx).font = Font(name="Arial", size=11, bold=True)
        worksheet.cell(total_row, col_idx).fill = PatternFill(fill_type="solid", fgColor="FFE6ECF4")

    worksheet.freeze_panes = "B9"


def write_summary_workbook(rows: list[StoreSalesReport], output_path: Path) -> None:
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Store Sales Summary"

    headers = ["Store Name", "Source File", "Date From", "Date To", "Total Sales"]
    worksheet.append(headers)

    header_fill = PatternFill(fill_type="solid", fgColor="FF1B365D")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFFFF")
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    total_sales_all_stores = 0.0
    for report in rows:
        worksheet.append([
            report.store_name,
            report.source_file,
            report.period_start,
            report.period_end,
            report.total_sales,
        ])
        total_sales_all_stores += report.total_sales

    total_row = worksheet.max_row + 1
    worksheet.cell(total_row, 1, "TOTAL")
    worksheet.cell(total_row, 5, total_sales_all_stores)

    for row_idx in range(2, total_row + 1):
        worksheet.cell(row_idx, 3).number_format = "dd-mmm-yyyy"
        worksheet.cell(row_idx, 4).number_format = "dd-mmm-yyyy"
        worksheet.cell(row_idx, 5).number_format = NUMFMT

    for cell in worksheet[total_row]:
        cell.font = Font(name="Arial", size=11, bold=True)
        cell.fill = PatternFill(fill_type="solid", fgColor="FFE6ECF4")

    worksheet.column_dimensions["A"].width = 24
    worksheet.column_dimensions["B"].width = 34
    worksheet.column_dimensions["C"].width = 16
    worksheet.column_dimensions["D"].width = 16
    worksheet.column_dimensions["E"].width = 16
    worksheet.freeze_panes = "A2"

    workbook.save(output_path)


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Calculate total sales for each store workbook and the overall total."
    )
    parser.add_argument(
        "files",
        nargs="*",
        help="Excel files to process. If omitted, all .xlsx files in the current directory are used.",
    )
    parser.add_argument(
        "-o",
        "--output",
        default="store_sales_summary.xlsx",
        help="Name of the summary workbook to create.",
    )
    args = parser.parse_args()

    output_name = Path(args.output).name
    paths = collect_files(args.files, output_name)
    if not paths:
        print("No Excel files found.")
        return 1

    results: list[StoreSalesReport] = []
    failed: list[tuple[str, str]] = []

    for path in paths:
        try:
            report = extract_store_report(path)
        except Exception as exc:
            failed.append((path.name, str(exc)))
            continue

        results.append(report)
        if report.period_start and report.period_end:
            period_text = f"{report.period_start.strftime('%d %b %Y')} to {report.period_end.strftime('%d %b %Y')}"
        else:
            period_text = "date range unavailable"
        print(f"{report.store_name}: ₹{report.total_sales:,.2f} ({period_text})")

    if not results:
        print("No files could be processed.")
        for filename, reason in failed:
            print(f"Skipped {filename}: {reason}")
        return 1

    grand_total = sum(report.total_sales for report in results)
    print(f"Grand Total: ₹{grand_total:,.2f}")

    output_path = Path(args.output).expanduser().resolve()
    write_summary_workbook(results, output_path)
    print(f"Saved summary workbook: {output_path}")

    if failed:
        print("\nSkipped files:")
        for filename, reason in failed:
            print(f"- {filename}: {reason}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
