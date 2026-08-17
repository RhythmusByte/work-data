# !pip install -q openpyxl pandas python-dateutil

import re
import pandas as pd
from datetime import datetime
from dateutil import parser as dateparser
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Upload the expenses_report_*.xlsx file (Date, Expense, Amount, Notes columns)
try:
    from google.colab import files
    uploaded = files.upload()
    uploaded_paths = list(uploaded.keys())
except ImportError:
    import glob
    uploaded_paths = glob.glob("expenses_report*.xlsx")

print("Files received:", uploaded_paths)

def find_header_row(ws, must_contain=("Date", "Expense", "Amount")):
    for r in range(1, min(ws.max_row, 20) + 1):
        row_vals = [ws.cell(r, c).value for c in range(1, 6)]
        row_str = [str(v).strip().lower() if v is not None else "" for v in row_vals]
        if all(tok.lower() in row_str for tok in must_contain):
            return r
    raise ValueError(f"Could not find a header row containing {must_contain}.")


def to_number(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return val
    s = str(val).replace(",", "").strip()
    try:
        return float(s)
    except ValueError:
        return val


ISO_DATE_RE = re.compile(r"^\d{4}-\d{2}-\d{2}$")

def parse_any_date(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v
    s = str(v).strip()
    if ISO_DATE_RE.match(s):
        try:
            return datetime.strptime(s, "%Y-%m-%d")
        except ValueError:
            return None
    try:
        return dateparser.parse(s, dayfirst=True)
    except (ValueError, TypeError):
        return None


# ---------------------------------------------------------------------------
# 2. Load every expense row from every uploaded file
# ---------------------------------------------------------------------------

expense_rows = []  # [date, category, amount, notes]

for path in uploaded_paths:
    wb_in = openpyxl.load_workbook(path, data_only=True)
    ws = wb_in[wb_in.sheetnames[0]]
    header_row = find_header_row(ws)
    print(f"{path}  ->  header at row {header_row}")
    for r in range(header_row + 1, ws.max_row + 1):
        date, category, amount, notes = (ws.cell(r, c).value for c in range(1, 5))
        if date is None and category is None:
            continue
        expense_rows.append([date, category, to_number(amount), notes])

print(f"\nTotal expense rows loaded: {len(expense_rows)}")

if not expense_rows:
    raise SystemExit("No expense rows found, check the uploaded file's layout.")

all_dates = [parse_any_date(r[0]) for r in expense_rows]
all_dates = [d for d in all_dates if d is not None]
period_start, period_end = min(all_dates), max(all_dates)
period_label = f"{period_start.strftime('%d %B %Y')} to {period_end.strftime('%d %B %Y')}"
print("Detected period:", period_label)

# Category list, in the order first seen (Summary sheet will re-sort by amount)
seen = []
for r in expense_rows:
    cat = (r[1] or "Uncategorised").strip()
    if cat not in seen:
        seen.append(cat)
print(f"Distinct expense categories found: {len(seen)}")
for c in seen:
    print("  -", c)
print("\n>>> Check this list. If two rows are really the same category spelled "
      "differently (e.g. 'Food Expense' vs 'Food Expenses'), fix the spelling in "
      "the source file and re-run, rather than editing the output by hand - "
      "otherwise next month's file will split them again.")

# ---------------------------------------------------------------------------
# 3. Build the workbook: Summary (main sheet) + Details
# ---------------------------------------------------------------------------

wb = openpyxl.Workbook()
ws_sum = wb.active
ws_sum.title = "Summary"
ws_det = wb.create_sheet("Expense Details")

FONT_NAME = "Arial"
NAVY = "FF1B365D"
GREY_FILL = "FFE6ECF4"
NUMFMT = "#,##0.00"

# --- Details sheet: raw rows, sorted by category then date -----------------
det_headers = ["Date", "Expense Category", "Amount", "Notes"]
ws_det.append(det_headers)
for cell in ws_det[1]:
    cell.font = Font(name=FONT_NAME, size=11, bold=True)

rows_sorted = sorted(
    expense_rows,
    key=lambda r: ((r[1] or "Uncategorised").strip(), parse_any_date(r[0]) or datetime.min),
)
for r in rows_sorted:
    date_val = parse_any_date(r[0])
    ws_det.append([date_val, (r[1] or "Uncategorised").strip(), r[2], r[3]])

for row in ws_det.iter_rows(min_row=2, min_col=1, max_col=1):
    row[0].number_format = "dd-mmm-yyyy"
for row in ws_det.iter_rows(min_row=2, min_col=3, max_col=3):
    row[0].number_format = NUMFMT

for col_idx, header in enumerate(det_headers, start=1):
    ws_det.column_dimensions[get_column_letter(col_idx)].width = max(16, len(header) + 4)
ws_det.freeze_panes = "A2"

det_last_row = ws_det.max_row  # last data row in Expense Details

# --- Summary sheet ----------------------------------------------------------
ws_sum.column_dimensions["B"].width = 34
ws_sum.column_dimensions["C"].width = 16
ws_sum.column_dimensions["D"].width = 12
ws_sum.column_dimensions["E"].width = 14

ws_sum["B2"] = "Expense Summary"
ws_sum["B2"].font = Font(name=FONT_NAME, size=14, bold=True, color=NAVY)
ws_sum["B3"] = f"Period: {period_label}"
ws_sum["B3"].font = Font(name=FONT_NAME, size=10, color="FF555555")

header_row_n = 5
headers = ["Expense Category", "Total Amount (INR)", "No. of Entries", "% of Total"]
for i, h in enumerate(headers, start=2):
    c = ws_sum.cell(header_row_n, i, h)
    c.font = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFFFF")
    c.fill = PatternFill(fill_type="solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="center")

# order categories by total amount, descending, computed once in Python purely
# to decide ROW ORDER - every number shown is still a live formula, not a
# hardcoded value, so the sheet recalculates if Expense Details changes.
totals_for_ordering = {}
for r in expense_rows:
    cat = (r[1] or "Uncategorised").strip()
    totals_for_ordering[cat] = totals_for_ordering.get(cat, 0) + (r[2] or 0)
categories_ordered = sorted(totals_for_ordering, key=lambda c: -totals_for_ordering[c])

row = header_row_n + 1
first_data_row = row
for cat in categories_ordered:
    ws_sum.cell(row, 2, cat).font = Font(name=FONT_NAME, size=11)
    amt_cell = ws_sum.cell(
        row, 3,
        f"=SUMIF('Expense Details'!$B$2:$B${det_last_row},$B{row},"
        f"'Expense Details'!$C$2:$C${det_last_row})"
    )
    amt_cell.number_format = NUMFMT
    cnt_cell = ws_sum.cell(
        row, 4,
        f"=COUNTIF('Expense Details'!$B$2:$B${det_last_row},$B{row})"
    )
    pct_cell = ws_sum.cell(row, 5, f"=C{row}/$C${first_data_row + len(categories_ordered)}")
    pct_cell.number_format = "0.0%"
    row += 1

total_row = row
ws_sum.cell(total_row, 2, "TOTAL").font = Font(name=FONT_NAME, size=11, bold=True)
tot_amt = ws_sum.cell(total_row, 3, f"=SUM(C{first_data_row}:C{total_row - 1})")
tot_amt.number_format = NUMFMT
tot_amt.font = Font(name=FONT_NAME, size=11, bold=True)
tot_cnt = ws_sum.cell(total_row, 4, f"=SUM(D{first_data_row}:D{total_row - 1})")
tot_cnt.font = Font(name=FONT_NAME, size=11, bold=True)
tot_pct = ws_sum.cell(total_row, 5, f"=SUM(E{first_data_row}:E{total_row - 1})")
tot_pct.number_format = "0.0%"
tot_pct.font = Font(name=FONT_NAME, size=11, bold=True)
for c in range(2, 6):
    ws_sum.cell(total_row, c).fill = PatternFill(fill_type="solid", fgColor=GREY_FILL)

# cross-check note: this should equal the total salary figure printed near the
# top of the raw report if the source report provides one.
ws_sum.cell(total_row + 2, 2, "Cross-check this TOTAL against the salary figure "
            "printed near the top of the raw expense report file.")
ws_sum.cell(total_row + 2, 2).font = Font(name=FONT_NAME, size=9, italic=True, color="FF888888")

out_name = f"Expense_Summary_{period_start.strftime('%Y-%m-%d')}_to_{period_end.strftime('%Y-%m-%d')}.xlsx"
wb.save(out_name)
print("\nSaved:", out_name)

try:
    from google.colab import files as colab_files
    colab_files.download(out_name)
except ImportError:
    pass
