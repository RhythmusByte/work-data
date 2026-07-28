# !pip install -q openpyxl pandas python-dateutil

import re
import pandas as pd
from datetime import datetime
from dateutil import parser as dateparser
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter

# sales_report_*.xlsx, expenses_report_*.xlsx, vendor_purchase_report_*.xlsx
try:
    from google.colab import files
    uploaded = files.upload()
    uploaded_paths = list(uploaded.keys())
except ImportError:
    import glob
    uploaded_paths = glob.glob("*.xlsx")

print("Files received:", uploaded_paths)


def find_header_row(ws, must_contain=("Date",)):
    for r in range(1, min(ws.max_row, 20) + 1):
        row_vals = [ws.cell(r, c).value for c in range(1, 6)]
        row_str = [str(v).strip() if v is not None else "" for v in row_vals]
        if all(any(tok.lower() == cell.lower() for cell in row_str) for tok in must_contain):
            return r
    raise ValueError(f"Could not find a header row containing {must_contain} "
                      f"in the first 20 rows of sheet '{ws.title}'.")


def to_number(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return val
    s = str(val).replace(",", "").strip()
    try:
        return float(s)
    except ValueError:
        return val 


def load_raw_sheet(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    return wb[wb.sheetnames[0]]


def identify_report(ws):
    header_row = find_header_row(ws, must_contain=("Date",))
    last_col = min(ws.max_column, 10)
    headers = [str(ws.cell(header_row, c).value or "").strip().lower()
               for c in range(1, last_col + 1)]
    if "payment method" in headers:
        return "sales", header_row
    if "vendor" in headers:
        return "purchase", header_row
    if "expense" in headers or "expense category" in headers:
        return "expense", header_row
    if "amount" in headers:
        return "purchase_return", header_row
    raise ValueError("Unrecognised report layout - check the file's column headers.")

sales_rows, purchase_rows, expense_rows, purchase_return_rows = [], [], [], []

for path in uploaded_paths:
    ws = load_raw_sheet(path)
    try:
        kind, header_row = identify_report(ws)
    except ValueError as e:
        print(f"{path}  ->  SKIPPED, not a recognised report ({e})")
        continue
    print(f"{path}  ->  detected as '{kind}' report (header at row {header_row})")

    if kind == "sales":
        for r in range(header_row + 1, ws.max_row + 1):
            row = [ws.cell(r, c).value for c in range(1, 9)]
            if row[0] is None and row[1] is None:
                continue
            sales_rows.append(row)

    elif kind == "purchase":
        for r in range(header_row + 1, ws.max_row + 1):
            row = [ws.cell(r, c).value for c in range(1, 6)]
            if row[0] is None and row[1] is None:
                continue
            row[3] = to_number(row[3])  
            purchase_rows.append(row)

    elif kind == "expense":
        for r in range(header_row + 1, ws.max_row + 1):
            row = [ws.cell(r, c).value for c in range(1, 5)]
            if row[0] is None and row[1] is None:
                continue
            expense_rows.append(row)

    elif kind == "purchase_return":
        for r in range(header_row + 1, ws.max_row + 1):
            row = [ws.cell(r, c).value for c in range(1, 3)]
            if row[0] is None and row[1] is None:
                continue
            row[1] = to_number(row[1])  
            purchase_return_rows.append(row)

print(f"\nSales rows: {len(sales_rows)} | Purchase rows: {len(purchase_rows)} | "
      f"Expense rows: {len(expense_rows)} | Purchase Return rows: {len(purchase_return_rows)}")

ISO_DATE_RE = re.compile(r"^\d{4}-\d{2}-\d{2}$")

def parse_any_date(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v
    s = str(v).strip()
    if ISO_DATE_RE.match(s):
        try:
            return datetime.strptime(s, "%Y-%m-%d")
        except ValueError:
            return None
    try:
        return dateparser.parse(s, dayfirst=True)
    except (ValueError, TypeError):
        return None

all_dates = [parse_any_date(r[0]) for r in sales_rows + purchase_rows + expense_rows + purchase_return_rows]
all_dates = [d for d in all_dates if d is not None]
period_start, period_end = min(all_dates), max(all_dates)
period_label = f"{period_start.strftime('%d %B %Y')} to {period_end.strftime('%d %B %Y')}"
print("Detected period:", period_label)

TOOLS_ACCESSORIES_EXACT = {"cleaning items", "cutting tools", "packing items"}

def classify_category(name):
    n = (name or "").strip().lower()
    if "salary" in n:
        return "salary"
    if n in TOOLS_ACCESSORIES_EXACT or "tool" in n or "accessor" in n:
        return "tools_accessories"
    return "standalone"

categories = sorted({row[1] for row in expense_rows if row[1]})
salary_categories = [c for c in categories if classify_category(c) == "salary"]
tools_categories = [c for c in categories if classify_category(c) == "tools_accessories"]
standalone_categories = [c for c in categories if classify_category(c) == "standalone"]

print("Merged into Total Salaries:      ", salary_categories)
print("Merged into Tools & Accessories: ", tools_categories)
print("Standalone expense lines:        ", standalone_categories)
print("\n>>> Check the 3 lists above before trusting the output. "
      "If a category landed in the wrong bucket, edit classify_category() and re-run from Cell 6.")

KNOWN_PAYMENT_METHODS = ["cash", "upi", "cash,upi", "credit"]
PAYMENT_LABELS = {
    "cash": "Cash Sales",
    "upi": "UPI Sales",
    "cash,upi": "Split Payment (Cash + UPI)",
    "credit": "Credit Sales (Pending Collection)",
}

payment_methods_seen = sorted({(row[4] or "").strip().lower() for row in sales_rows})
unexpected_methods = [p for p in payment_methods_seen if p not in KNOWN_PAYMENT_METHODS]
if unexpected_methods:
    print("!! New payment method(s) not seen before, will fall into 'Other / Unclassified Sales':",
          unexpected_methods)
else:
    print("Payment methods this period:", payment_methods_seen, "(all recognised)")

wb = openpyxl.Workbook()

ws_pnl = wb.active
ws_pnl.title = "P&L Summary"
ws_sales = wb.create_sheet("Sales Details")
ws_purchase = wb.create_sheet("Purchase Details")
ws_expense = wb.create_sheet("Expense Details")
ws_purchase_return = wb.create_sheet("Purchase Return Details")

SALES_HEADERS = ["Date", "Bill No", "Customer Name", "Customer Mobile No",
                  "Payment Method", "Sales Via", "Coupon Discount", "Total Rate (INR)"]
PURCHASE_HEADERS = ["Date", "Vendor", "Invoice", "Amount", "Status"]
EXPENSE_HEADERS = ["Date", "Expense Category", "Amount", "Notes"]
PURCHASE_RETURN_HEADERS = ["Date", "Amount"]

def write_sheet(ws, headers, rows):
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(name="Segoe UI", size=11, bold=True)
    for row in rows:
        ws.append(row)
    for col_idx, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(14, len(header) + 2)
    ws.freeze_panes = "A2"

write_sheet(ws_sales, SALES_HEADERS, sales_rows)
write_sheet(ws_purchase, PURCHASE_HEADERS, purchase_rows)
write_sheet(ws_expense, EXPENSE_HEADERS, expense_rows)
write_sheet(ws_purchase_return, PURCHASE_RETURN_HEADERS, purchase_return_rows)

n_sales = len(sales_rows) 
n_purchase = len(purchase_rows)
n_expense = len(expense_rows)
n_purchase_return = len(purchase_return_rows)

sales_last_row = n_sales + 1
purchase_last_row = n_purchase + 1
expense_last_row = n_expense + 1
purchase_return_last_row = n_purchase_return + 1

NUMFMT = "#,##0.00"
NAVY = "FF1B365D"
GREY_FILL = "FFE6ECF4"

ws_pnl.column_dimensions["B"].width = 40
ws_pnl.column_dimensions["C"].width = 20

def set_row(r, label, formula=None, bold=False, fill=False, section=False, size=11, comment=None):
    bcell, ccell = ws_pnl.cell(r, 2), ws_pnl.cell(r, 3)
    bcell.value = label
    color = NAVY if section else ("FF000000" if not bold else "FF000000")
    f = Font(name="Segoe UI", size=size, bold=bold, color=color)
    bcell.font, ccell.font = f, f
    if fill:
        pf = PatternFill(fill_type="solid", fgColor=GREY_FILL)
        bcell.fill, ccell.fill = pf, pf
    if formula is not None:
        ccell.value = formula
        ccell.number_format = NUMFMT
    if comment:
        bcell.comment = Comment(comment, "Claude")

ws_pnl["B3"] = "Profit & Loss Statement"
ws_pnl["B3"].font = Font(name="Segoe UI", size=12, bold=True, color="FF555555")
ws_pnl["B4"] = f"Period: {period_label}"
ws_pnl["B4"].font = Font(name="Segoe UI", size=10, color="FF555555")

row = 7

set_row(row, "Revenue", section=True, bold=True); row += 1
cash_row = row
set_row(row, "  Cash Sales", f"=SUMIF('Sales Details'!E2:E{sales_last_row}, \"cash\", 'Sales Details'!H2:H{sales_last_row})"); row += 1
upi_row = row
set_row(row, "  UPI Sales", f"=SUMIF('Sales Details'!E2:E{sales_last_row}, \"upi\", 'Sales Details'!H2:H{sales_last_row})"); row += 1
split_row = row
set_row(row, "  Split Payment (Cash + UPI)",
        f"=SUMIF('Sales Details'!E2:E{sales_last_row}, \"cash,upi\", 'Sales Details'!H2:H{sales_last_row})",
        comment="Bills where the POS recorded payment as 'cash,upi' with no split amount. "
                "Kept separate rather than guessed into Cash or UPI.")
row += 1
credit_row = row
set_row(row, "  Credit Sales (Pending Collection)",
        f"=SUMIF('Sales Details'!E2:E{sales_last_row}, \"credit\", 'Sales Details'!H2:H{sales_last_row})",
        comment="Revenue recognised on an accrual basis for bills sold on credit; not yet collected as cash/UPI.")
row += 1
other_rev_row = row
set_row(row, "  Other / Unclassified Sales",
        f"=SUM('Sales Details'!H2:H{sales_last_row})-SUM(C{cash_row}:C{credit_row})",
        comment="Balancing line: Total Revenue minus the 4 known payment methods above. "
                "Should be 0.00 - if it isn't, a new payment method type showed up this period; check Sales Details column E.")
row += 1
total_rev_row = row
set_row(row, "Total Revenue", f"=SUM(C{cash_row}:C{other_rev_row})", bold=True, fill=True)
row += 2

set_row(row, "Cost of Goods Sold (COGS)", section=True, bold=True); row += 1
gross_purchase_row = row
set_row(row, "  Vendor Purchases (Gross)",
        f"=SUMIF('Purchase Details'!E2:E{purchase_last_row}, \"<>Return\", 'Purchase Details'!D2:D{purchase_last_row})",
        comment="All Purchase Details rows EXCEPT Status = 'Return'. Matches 'anything that isn't a "
                "Return' rather than a specific status word like 'Unpaid', so this keeps working once "
                "purchases start getting marked 'Paid' too.")
row += 1
vendor_return_row = row
set_row(row, "  Less: Purchase Return (in Vendor Report)",
        f"=SUMIF('Purchase Details'!E2:E{purchase_last_row}, \"Return\", 'Purchase Details'!D2:D{purchase_last_row})",
        comment="Covers the case where a return is entered as its own row inside the vendor purchase "
                "report, Status = 'Return', NEGATIVE amount. Reads 0.00 if returns are only tracked "
                "via the separate Purchase Return report instead (see next line).")
row += 1
separate_return_row = row
set_row(row, "  Less: Purchase Return (separate report)",
        f"=-SUM('Purchase Return Details'!B2:B{purchase_return_last_row})",
        comment="From the standalone Purchase Return file (Date + Amount columns, amounts entered as "
                "POSITIVE there). Negated here since it reduces COGS.")
row += 1
total_cogs_row = row
set_row(row, "Total COGS", f"=SUM(C{gross_purchase_row}:C{separate_return_row})", bold=True, fill=True)
row += 1
gross_profit_row = row
set_row(row, "Gross Profit", f"=C{total_rev_row}-C{total_cogs_row}", bold=True, fill=True)
row += 2

set_row(row, "Operating Expenses", section=True, bold=True); row += 1

salary_row = row
set_row(row, "  Total Salaries (All Roles)",
        f"=SUMPRODUCT(ISNUMBER(SEARCH(\"salary\",'Expense Details'!B2:B{expense_last_row}))"
        f"*'Expense Details'!C2:C{expense_last_row})",
        comment=f"Merges: {', '.join(salary_categories) if salary_categories else '(none found this period)'}. "
                "Text-search based, so a new '...Salary' category is picked up automatically.")
row += 1

tools_formula_parts = [f"SUMIF('Expense Details'!B2:B{expense_last_row}, \"{c}\", 'Expense Details'!C2:C{expense_last_row})"
                        for c in tools_categories]
tools_formula = "=" + "+".join(tools_formula_parts) if tools_formula_parts else "=0"
tools_row = row
set_row(row, "  Tools & Accessories", tools_formula,
        comment=f"Merges: {', '.join(tools_categories) if tools_categories else '(none found this period)'}.")
row += 1

for cat in standalone_categories:
    set_row(row, f"  {cat}",
            f"=SUMIF('Expense Details'!B2:B{expense_last_row}, \"{cat}\", 'Expense Details'!C2:C{expense_last_row})")
    row += 1

total_opex_row = row
set_row(total_opex_row, "Total Operating Expenses", f"=SUM(C{salary_row}:C{total_opex_row - 1})", bold=True, fill=True)
row += 2

net_row = row
set_row(net_row, "Net Profit / (Loss)", f"=C{gross_profit_row}-C{total_opex_row}", bold=True, fill=True, size=12)

out_name = f"PNL_{period_start.strftime('%Y-%m-%d')}_to_{period_end.strftime('%Y-%m-%d')}.xlsx"
wb.save(out_name)
print("Saved:", out_name)

try:
    from google.colab import files as colab_files
    colab_files.download(out_name)
except ImportError:
    pass