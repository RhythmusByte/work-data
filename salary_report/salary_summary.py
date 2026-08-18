from __future__ import annotations

import argparse
import glob
import re
import sys
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

import openpyxl
from dateutil import parser as dateparser
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


SCRIPT_DIR = Path(__file__).resolve().parent
HEADER_SCAN_LIMIT = 20
MAX_PREVIEW_COLUMNS = 6
NUMFMT = "#,##0.00"
TITLE_FILL = "FF1B365D"
LIGHT_FILL = "FFE6ECF4"
OUTPUT_PREFIX = "Salary_Summary_"
EXPECTED_HEADERS = ("Date", "Expense", "Amount", "Notes")

ISO_DATE_RE = re.compile(r"^\d{4}-\d{2}-\d{2}$")
INVALID_SHEET_CHARS = re.compile(r"[\[\]\:\*\?\/\\]")


@dataclass
class StoreReport:
    store_name: str
    source_file: str
    total_salary: float
    salary_rows: list[list[object]]
    period_start: datetime | None
    period_end: datetime | None
    sheet_name: str
    category_totals: dict[str, list] = field(default_factory=dict)
    month_totals: dict[str, list] = field(default_factory=dict)
    month_labels: dict[str, str] = field(default_factory=dict)
    food_rows: list[list[object]] = field(default_factory=list)
    food_total: float = 0.0
    food_category_totals: dict[str, list] = field(default_factory=dict)
    food_month_totals: dict[str, list] = field(default_factory=dict)
    food_month_labels: dict[str, str] = field(default_factory=dict)


def normalize_text(value) -> str:
    return str(value or "").strip().lower()


def to_number(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace(",", "").strip()
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def parse_any_date(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    text = str(value).strip()
    if not text:
        return None
    if ISO_DATE_RE.match(text):
        try:
            return datetime.strptime(text, "%Y-%m-%d")
        except ValueError:
            return None
    try:
        return dateparser.parse(text, dayfirst=True)
    except (ValueError, TypeError):
        return None


def find_header_row(ws) -> int:
    expected = [normalize_text(h) for h in EXPECTED_HEADERS]
    for row_idx in range(1, min(ws.max_row, HEADER_SCAN_LIMIT) + 1):
        values = [
            normalize_text(ws.cell(row_idx, col_idx).value)
            for col_idx in range(1, min(ws.max_column, MAX_PREVIEW_COLUMNS) + 1)
        ]
        if values[: len(expected)] == expected:
            return row_idx
    raise ValueError(f"Could not find a header row matching {EXPECTED_HEADERS}.")


def is_salary_row(row) -> bool:
    category = normalize_text(row[1])
    notes = normalize_text(row[3])
    return "salary" in category or "salary" in notes


def is_food_row(row) -> bool:
    category = normalize_text(row[1])
    notes = normalize_text(row[3])
    return "food" in category or "food" in notes


def collect_files(file_args: list[str], output_name: str) -> list[Path]:
    if file_args:
        paths = [Path(arg).expanduser().resolve() for arg in file_args]
    else:
        try:
            from google.colab import files

            uploaded = files.upload()
            paths = [Path(name).resolve() for name in uploaded.keys()]
        except ImportError:
            paths = [Path(path).resolve() for path in glob.glob(str(SCRIPT_DIR / "*.xlsx"))]
            if not paths:
                paths = [Path(path).resolve() for path in glob.glob("*.xlsx")]

    result = []
    for path in paths:
        if not path.is_file():
            continue
        if path.suffix.lower() != ".xlsx":
            continue
        if path.name.startswith("~$"):
            continue
        if path.name.startswith(OUTPUT_PREFIX):
            continue
        if path.name == output_name:
            continue
        result.append(path)

    return sorted(result)


def prompt_store_names(paths: list[Path]) -> dict[Path, str]:
    print("Files received:")
    for path in paths:
        print(f"  - {path.name}")

    interactive = sys.stdin.isatty()
    used_names: set[str] = set()
    store_names: dict[Path, str] = {}

    for path in paths:
        default_name = path.stem.replace("_", " ").replace("-", " ").strip() or "Store"
        if interactive:
            while True:
                response = input(f"Store name for file {path.name} [{default_name}]: ").strip()
                candidate = " ".join((response or default_name).split())
                if candidate not in used_names:
                    break
                print(f"Store name '{candidate}' is already used. Choose a unique name.")
        else:
            candidate = default_name if default_name not in used_names else f"{default_name} {len(used_names) + 1}"
        used_names.add(candidate)
        store_names[path] = candidate

    return store_names


def make_unique_sheet_title(base_name: str, existing: set[str]) -> str:
    cleaned = INVALID_SHEET_CHARS.sub(" ", base_name).strip()
    cleaned = re.sub(r"\s+", " ", cleaned) or "Store"
    suffix = " Summary"
    max_base_len = 31 - len(suffix)
    base = cleaned[:max_base_len].rstrip()
    candidate = f"{base}{suffix}"
    counter = 2
    while candidate in existing:
        extra = f" ({counter})"
        max_base_len = 31 - len(suffix) - len(extra)
        base = cleaned[:max_base_len].rstrip()
        candidate = f"{base}{suffix}{extra}"
        counter += 1
    existing.add(candidate)
    return candidate


def extract_store_report(path: Path, store_name: str, sheet_name: str) -> StoreReport:
    workbook = openpyxl.load_workbook(path, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    header_row = find_header_row(worksheet)

    salary_rows: list[list[object]] = []
    food_rows: list[list[object]] = []
    for row_idx in range(header_row + 1, worksheet.max_row + 1):
        row = [worksheet.cell(row_idx, col_idx).value for col_idx in range(1, 5)]
        if row[0] is None and row[1] is None and row[2] is None and row[3] is None:
            continue
        parsed_row = [row[0], row[1], to_number(row[2]), row[3]]
        if is_salary_row(row):
            salary_rows.append(parsed_row)
        elif is_food_row(row):
            food_rows.append(parsed_row)

    total_salary = sum((row[2] or 0) for row in salary_rows)
    dates = [parse_any_date(row[0]) for row in salary_rows]
    dates = [dt for dt in dates if dt is not None]
    period_start = min(dates) if dates else None
    period_end = max(dates) if dates else None

    category_totals: dict[str, list] = {}
    month_totals: dict[str, list] = {}
    month_labels: dict[str, str] = {}
    for row in salary_rows:
        category_name = str(row[1] or "").strip() or "Uncategorised"
        cat_entry = category_totals.setdefault(category_name, [0, 0.0])
        cat_entry[0] += 1
        cat_entry[1] += row[2] or 0

        row_date = parse_any_date(row[0])
        month_key = row_date.strftime("%Y-%m") if row_date else "Unknown"
        month_labels[month_key] = row_date.strftime("%B %Y") if row_date else "Unknown"
        month_entry = month_totals.setdefault(month_key, [0, 0.0])
        month_entry[0] += 1
        month_entry[1] += row[2] or 0

    food_category_totals: dict[str, list] = {}
    food_month_totals: dict[str, list] = {}
    food_month_labels: dict[str, str] = {}
    for row in food_rows:
        category_name = str(row[1] or "").strip() or "Uncategorised"
        cat_entry = food_category_totals.setdefault(category_name, [0, 0.0])
        cat_entry[0] += 1
        cat_entry[1] += row[2] or 0

        row_date = parse_any_date(row[0])
        month_key = row_date.strftime("%Y-%m") if row_date else "Unknown"
        food_month_labels[month_key] = row_date.strftime("%B %Y") if row_date else "Unknown"
        month_entry = food_month_totals.setdefault(month_key, [0, 0.0])
        month_entry[0] += 1
        month_entry[1] += row[2] or 0

    food_total = sum((row[2] or 0) for row in food_rows)

    return StoreReport(
        store_name=store_name,
        source_file=path.name,
        total_salary=total_salary,
        salary_rows=salary_rows,
        period_start=period_start,
        period_end=period_end,
        sheet_name=sheet_name,
        category_totals=category_totals,
        month_totals=month_totals,
        month_labels=month_labels,
        food_rows=food_rows,
        food_total=food_total,
        food_category_totals=food_category_totals,
        food_month_totals=food_month_totals,
        food_month_labels=food_month_labels,
    )


def _write_section_header(ws, row_idx: int, text: str) -> None:
    cell = ws.cell(row_idx, 2, text)
    cell.font = Font(name="Arial", size=11, bold=True, color=TITLE_FILL)


def _write_table_header(ws, row_idx: int, headers: list[str], start_col: int = 2) -> None:
    for offset, header in enumerate(headers):
        cell = ws.cell(row_idx, start_col + offset, header)
        cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFFFF")
        cell.fill = PatternFill(fill_type="solid", fgColor=TITLE_FILL)
        cell.alignment = Alignment(horizontal="center")


def _write_total_row(ws, row_idx: int, start_col: int, end_col: int) -> None:
    for col_idx in range(start_col, end_col + 1):
        ws.cell(row_idx, col_idx).font = Font(name="Arial", size=10, bold=True)
        ws.cell(row_idx, col_idx).fill = PatternFill(fill_type="solid", fgColor=LIGHT_FILL)


def _write_monthly_breakdown(ws, row: int, title: str, month_totals: dict, month_labels: dict, rows_list: list, grand_total: float, no_rows_text: str) -> int:
    _write_section_header(ws, row, title)
    row += 1
    _write_table_header(ws, row, ["Month", "Rows", "Total (INR)"])
    row += 1
    sorted_months = sorted(month_totals.items(), key=lambda item: item[0])
    for month_key, (rows_count, month_total) in sorted_months:
        ws.cell(row, 2, month_labels.get(month_key, month_key))
        ws.cell(row, 3, rows_count)
        ws.cell(row, 4, month_total)
        ws.cell(row, 4).number_format = NUMFMT
        row += 1
    if sorted_months:
        ws.cell(row, 2, "TOTAL")
        ws.cell(row, 3, len(rows_list))
        ws.cell(row, 4, grand_total)
        ws.cell(row, 4).number_format = NUMFMT
        _write_total_row(ws, row, 2, 4)
        row += 1
    else:
        ws.cell(row, 2, no_rows_text)
        ws.cell(row, 2).font = Font(name="Arial", size=10, italic=True, color="FF888888")
        row += 1
    return row + 1


def _write_category_breakdown(ws, row: int, title: str, category_totals: dict, rows_list: list, grand_total: float, no_rows_text: str) -> int:
    _write_section_header(ws, row, title)
    row += 1
    _write_table_header(ws, row, ["Category", "Rows", "Total (INR)"])
    row += 1
    sorted_categories = sorted(category_totals.items(), key=lambda item: item[1][1], reverse=True)
    for category_name, (rows_count, category_total) in sorted_categories:
        ws.cell(row, 2, category_name)
        ws.cell(row, 3, rows_count)
        ws.cell(row, 4, category_total)
        ws.cell(row, 4).number_format = NUMFMT
        row += 1
    if sorted_categories:
        ws.cell(row, 2, "TOTAL")
        ws.cell(row, 3, len(rows_list))
        ws.cell(row, 4, grand_total)
        ws.cell(row, 4).number_format = NUMFMT
        _write_total_row(ws, row, 2, 4)
        row += 1
    else:
        ws.cell(row, 2, no_rows_text)
        ws.cell(row, 2).font = Font(name="Arial", size=10, italic=True, color="FF888888")
        row += 1
    return row + 1


def _write_detail_table(ws, row: int, title: str, rows_list: list, grand_total: float, no_rows_text: str) -> int:
    _write_section_header(ws, row, title)
    row += 1
    header_row = row
    _write_table_header(ws, header_row, ["Date", "Expense Category", "Amount", "Notes"])

    for row_idx, data_row in enumerate(rows_list, start=header_row + 1):
        ws.cell(row_idx, 2, parse_any_date(data_row[0]) or data_row[0])
        ws.cell(row_idx, 3, data_row[1] or "Uncategorised")
        ws.cell(row_idx, 4, data_row[2])
        ws.cell(row_idx, 5, data_row[3])

    for r in ws.iter_rows(min_row=header_row + 1, min_col=2, max_col=2, max_row=header_row + len(rows_list)):
        r[0].number_format = "dd-mmm-yyyy"
    for r in ws.iter_rows(min_row=header_row + 1, min_col=4, max_col=4, max_row=header_row + len(rows_list)):
        r[0].number_format = NUMFMT

    if rows_list:
        total_row = header_row + len(rows_list) + 1
        ws.cell(total_row, 3, "TOTAL")
        ws.cell(total_row, 4, grand_total)
        ws.cell(total_row, 4).number_format = NUMFMT
        _write_total_row(ws, total_row, 3, 5)
        return total_row + 2
    else:
        ws.cell(header_row + 1, 2, no_rows_text)
        ws.cell(header_row + 1, 2).font = Font(name="Arial", size=10, italic=True, color="FF888888")
        return header_row + 3


def write_store_sheet(ws, report: StoreReport) -> None:
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 18

    ws["B2"] = f"{report.store_name} Summary"
    ws["B2"].font = Font(name="Arial", size=14, bold=True, color=TITLE_FILL)
    ws["B3"] = f"Source file: {report.source_file}"
    ws["B3"].font = Font(name="Arial", size=10, color="FF555555")
    ws["B4"] = f"Salary rows matched: {len(report.salary_rows)} | Food rows matched: {len(report.food_rows)}"
    ws["B4"].font = Font(name="Arial", size=10, color="FF555555")
    period_text = "Period: not available"
    if report.period_start and report.period_end:
        period_text = f"Period: {report.period_start.strftime('%d %B %Y')} to {report.period_end.strftime('%d %B %Y')}"
    ws["B5"] = period_text
    ws["B5"].font = Font(name="Arial", size=10, color="FF555555")
    ws["B6"] = f"Total salary given (INR): {report.total_salary:,.2f}"
    ws["B6"].font = Font(name="Arial", size=11, bold=True)
    ws["B7"] = f"Total food expense (INR): {report.food_total:,.2f}"
    ws["B7"].font = Font(name="Arial", size=11, bold=True)

    row = 9

    row = _write_monthly_breakdown(
        ws, row, "Salary - Monthly Breakdown", report.month_totals, report.month_labels,
        report.salary_rows, report.total_salary, "No salary rows were found in this file.",
    )
    row = _write_category_breakdown(
        ws, row, "Salary - Category Breakdown", report.category_totals,
        report.salary_rows, report.total_salary, "No salary categories found.",
    )
    row = _write_monthly_breakdown(
        ws, row, "Food - Monthly Breakdown", report.food_month_totals, report.food_month_labels,
        report.food_rows, report.food_total, "No food rows were found in this file.",
    )
    row = _write_category_breakdown(
        ws, row, "Food - Category Breakdown", report.food_category_totals,
        report.food_rows, report.food_total, "No food categories found.",
    )

    row = _write_detail_table(
        ws, row, "Salary - Detail", report.salary_rows, report.total_salary,
        "No salary rows were found in this file.",
    )
    freeze_row = row
    row = _write_detail_table(
        ws, row, "Food - Detail", report.food_rows, report.food_total,
        "No food rows were found in this file.",
    )

    ws.freeze_panes = f"B9"


def _write_matrix(ws, row: int, section_title: str, reports: list[StoreReport], keys_in_order: list, key_labels: dict, get_entry, get_report_total) -> int:
    _write_section_header(ws, row, section_title)
    row += 1
    header_row = row
    ws.cell(header_row, 2, key_labels.get("__col1__", "Item"))
    for idx, report in enumerate(reports):
        ws.cell(header_row, 4 + idx, report.store_name)
    grand_col = 4 + len(reports)
    ws.cell(header_row, grand_col, "Grand Total (INR)")
    _write_table_header(ws, header_row, [ws.cell(header_row, c).value for c in range(2, grand_col + 1)])
    row += 1

    key_grand_totals: dict[str, float] = {}
    for key in keys_in_order:
        row_total = 0.0
        ws.cell(row, 2, key_labels.get(key, key))
        for idx, report in enumerate(reports):
            _, value = get_entry(report, key)
            ws.cell(row, 4 + idx, value)
            ws.cell(row, 4 + idx).number_format = NUMFMT
            row_total += value
        ws.cell(row, grand_col, row_total)
        ws.cell(row, grand_col).number_format = NUMFMT
        key_grand_totals[key] = row_total
        row += 1

    ws.cell(row, 2, "TOTAL")
    for idx, report in enumerate(reports):
        ws.cell(row, 4 + idx, get_report_total(report))
        ws.cell(row, 4 + idx).number_format = NUMFMT
    ws.cell(row, grand_col, sum(key_grand_totals.values()))
    ws.cell(row, grand_col).number_format = NUMFMT
    _write_total_row(ws, row, 2, grand_col)
    row += 1

    return row + 1


def write_combined_summary_sheet(ws, reports: list[StoreReport]) -> None:
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 30
    for idx in range(len(reports)):
        ws.column_dimensions[get_column_letter(4 + idx)].width = 18
    ws.column_dimensions[get_column_letter(4 + len(reports))].width = 18

    ws["B2"] = "Salary and Food Expense Summary"
    ws["B2"].font = Font(name="Arial", size=14, bold=True, color=TITLE_FILL)

    all_dates = [dt for report in reports for dt in (report.period_start, report.period_end) if dt is not None]
    if all_dates:
        period_label = f"{min(all_dates).strftime('%d %B %Y')} to {max(all_dates).strftime('%d %B %Y')}"
    else:
        period_label = "not available"
    ws["B3"] = f"Period: {period_label}"
    ws["B3"].font = Font(name="Arial", size=10, color="FF555555")

    row = 5

    # Store totals table (salary and food side by side)
    _write_section_header(ws, row, "Store Totals")
    row += 1
    _write_table_header(ws, row, ["Store Name", "Source File", "Salary Rows", "Salary Total (INR)", "Food Rows", "Food Total (INR)"])
    row += 1
    for report in reports:
        ws.cell(row, 2, report.store_name)
        ws.cell(row, 3, report.source_file)
        ws.cell(row, 4, len(report.salary_rows))
        ws.cell(row, 5, report.total_salary)
        ws.cell(row, 5).number_format = NUMFMT
        ws.cell(row, 6, len(report.food_rows))
        ws.cell(row, 7, report.food_total)
        ws.cell(row, 7).number_format = NUMFMT
        row += 1
    ws.cell(row, 2, "TOTAL")
    ws.cell(row, 4, sum(len(report.salary_rows) for report in reports))
    ws.cell(row, 5, sum(report.total_salary for report in reports))
    ws.cell(row, 5).number_format = NUMFMT
    ws.cell(row, 6, sum(len(report.food_rows) for report in reports))
    ws.cell(row, 7, sum(report.food_total for report in reports))
    ws.cell(row, 7).number_format = NUMFMT
    _write_total_row(ws, row, 2, 7)
    row += 2

    # Monthly totals - salary
    month_keys: list[str] = []
    seen_months: set[str] = set()
    month_labels_all: dict[str, str] = {}
    for report in reports:
        for month_key in report.month_totals:
            month_labels_all[month_key] = report.month_labels.get(month_key, month_key)
            if month_key not in seen_months:
                seen_months.add(month_key)
                month_keys.append(month_key)
    month_keys.sort()
    key_labels = dict(month_labels_all)
    key_labels["__col1__"] = "Month"
    row = _write_matrix(
        ws, row, "Salary - Monthly Totals", reports, month_keys, key_labels,
        lambda report, key: report.month_totals.get(key, (0, 0.0)),
        lambda report: report.total_salary,
    )

    # Monthly totals - food
    food_month_keys: list[str] = []
    seen_food_months: set[str] = set()
    food_month_labels_all: dict[str, str] = {}
    for report in reports:
        for month_key in report.food_month_totals:
            food_month_labels_all[month_key] = report.food_month_labels.get(month_key, month_key)
            if month_key not in seen_food_months:
                seen_food_months.add(month_key)
                food_month_keys.append(month_key)
    food_month_keys.sort()
    food_key_labels = dict(food_month_labels_all)
    food_key_labels["__col1__"] = "Month"
    row = _write_matrix(
        ws, row, "Food - Monthly Totals", reports, food_month_keys, food_key_labels,
        lambda report, key: report.food_month_totals.get(key, (0, 0.0)),
        lambda report: report.food_total,
    )

    # Category totals - salary
    category_names: list[str] = []
    seen_categories: set[str] = set()
    category_grand_totals: dict[str, float] = {}
    for report in reports:
        for category_name, (_, category_total) in report.category_totals.items():
            category_grand_totals[category_name] = category_grand_totals.get(category_name, 0.0) + category_total
            if category_name not in seen_categories:
                seen_categories.add(category_name)
                category_names.append(category_name)
    category_names.sort(key=lambda name: category_grand_totals[name], reverse=True)
    cat_key_labels = {name: name for name in category_names}
    cat_key_labels["__col1__"] = "Category"
    row = _write_matrix(
        ws, row, "Salary - Category Totals", reports, category_names, cat_key_labels,
        lambda report, key: report.category_totals.get(key, (0, 0.0)),
        lambda report: report.total_salary,
    )

    # Category totals - food
    food_category_names: list[str] = []
    seen_food_categories: set[str] = set()
    food_category_grand_totals: dict[str, float] = {}
    for report in reports:
        for category_name, (_, category_total) in report.food_category_totals.items():
            food_category_grand_totals[category_name] = food_category_grand_totals.get(category_name, 0.0) + category_total
            if category_name not in seen_food_categories:
                seen_food_categories.add(category_name)
                food_category_names.append(category_name)
    food_category_names.sort(key=lambda name: food_category_grand_totals[name], reverse=True)
    food_cat_key_labels = {name: name for name in food_category_names}
    food_cat_key_labels["__col1__"] = "Category"
    row = _write_matrix(
        ws, row, "Food - Category Totals", reports, food_category_names, food_cat_key_labels,
        lambda report, key: report.food_category_totals.get(key, (0, 0.0)),
        lambda report: report.food_total,
    )

    ws.freeze_panes = "D6"


def build_workbook(reports: list[StoreReport], output_path: Path) -> None:
    workbook = openpyxl.Workbook()
    summary_ws = workbook.active
    summary_ws.title = "Salary Summary"
    write_combined_summary_sheet(summary_ws, reports)

    for report in reports:
        ws = workbook.create_sheet(report.sheet_name)
        write_store_sheet(ws, report)

    workbook.save(output_path)


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Create a salary and food expense report with one summary sheet and one sheet per store."
    )
    parser.add_argument(
        "files",
        nargs="*",
        help="Excel files to process. If omitted, all .xlsx files in the salary_report directory are used.",
    )
    parser.add_argument(
        "-o",
        "--output",
        default=None,
        help="Optional output workbook name. Defaults to Salary_Summary_<period>.xlsx.",
    )
    args = parser.parse_args()

    output_name = Path(args.output).name if args.output else ""
    paths = collect_files(args.files, output_name)
    if not paths:
        print("No Excel files found.")
        return 1

    store_names = prompt_store_names(paths)
    sheet_names: set[str] = set()
    reports: list[StoreReport] = []

    for path in paths:
        store_name = store_names[path]
        sheet_name = make_unique_sheet_title(store_name, sheet_names)
        report = extract_store_report(path, store_name, sheet_name)
        reports.append(report)
        print(f"{store_name} ({path.name}): Salary {report.total_salary:,.2f} | Food {report.food_total:,.2f}")

    all_dates = [dt for report in reports for dt in (report.period_start, report.period_end) if dt is not None]
    if all_dates:
        period_start = min(all_dates)
        period_end = max(all_dates)
        default_output = f"{OUTPUT_PREFIX}{period_start.strftime('%Y-%m-%d')}_to_{period_end.strftime('%Y-%m-%d')}.xlsx"
    else:
        default_output = f"{OUTPUT_PREFIX}report.xlsx"

    output_path = Path(args.output or (SCRIPT_DIR / default_output)).expanduser().resolve()
    build_workbook(reports, output_path)

    grand_total_salary = sum(report.total_salary for report in reports)
    grand_total_food = sum(report.food_total for report in reports)
    print(f"Grand Total Salary: {grand_total_salary:,.2f}")
    print(f"Grand Total Food: {grand_total_food:,.2f}")
    print(f"Saved summary workbook: {output_path}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())